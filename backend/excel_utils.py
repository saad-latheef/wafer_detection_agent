"""
Excel export utilities for wafer analysis data.
"""
import openpyxl
from openpyxl.chart import PieChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
from datetime import datetime
from typing import Dict, List, Any


def create_wafer_report_excel(
    lot_data: Dict[str, Any],
    wafer_analyses: List[Dict[str, Any]],
    trends: List[Dict[str, Any]] = None
) -> BytesIO:
    """
    Create an Excel workbook with wafer analysis data.
    
    Returns:
        BytesIO buffer containing the Excel file
    """
    wb = openpyxl.Workbook()
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="00D4FF", end_color="00D4FF", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # === Summary Sheet ===
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Title
    ws_summary["A1"] = "Wafer Analysis Report"
    ws_summary["A1"].font = Font(bold=True, size=16)
    ws_summary["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    
    # Lot Statistics
    ws_summary["A4"] = "Lot Statistics"
    ws_summary["A4"].font = Font(bold=True, size=12)
    
    stats = [
        ["Metric", "Value"],
        ["Total Wafers", lot_data.get("total_wafers", 0)],
        ["Defective Wafers", lot_data.get("defective_wafers", 0)],
        ["Yield Rate", f"{lot_data.get('yield_rate', 0):.2f}%"],
    ]
    
    for row_idx, row in enumerate(stats, start=5):
        for col_idx, value in enumerate(row, start=1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            if row_idx == 5:
                cell.font = header_font
                cell.fill = header_fill
    
    # Defect Distribution
    if "defect_distribution" in lot_data:
        ws_summary["A11"] = "Defect Distribution"
        ws_summary["A11"].font = Font(bold=True, size=12)
        
        dist_headers = ["Pattern", "Count", "Percentage"]
        for col_idx, header in enumerate(dist_headers, start=1):
            cell = ws_summary.cell(row=12, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        total = lot_data.get("total_wafers", 1)
        for row_idx, (pattern, count) in enumerate(lot_data["defect_distribution"].items(), start=13):
            ws_summary.cell(row=row_idx, column=1, value=pattern).border = border
            ws_summary.cell(row=row_idx, column=2, value=count).border = border
            ws_summary.cell(row=row_idx, column=3, value=f"{count/total*100:.1f}%").border = border
    
    # Adjust column widths
    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 15
    ws_summary.column_dimensions["C"].width = 15
    
    # === Wafer Details Sheet ===
    ws_wafers = wb.create_sheet("Wafer Details")
    
    headers = ["Wafer ID", "File Name", "Verdict", "Confidence", "Severity", "Detected Pattern"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_wafers.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    for row_idx, wafer in enumerate(wafer_analyses, start=2):
        ws_wafers.cell(row=row_idx, column=1, value=wafer.get("waferId", "")).border = border
        ws_wafers.cell(row=row_idx, column=2, value=wafer.get("fileName", "")).border = border
        
        verdict_cell = ws_wafers.cell(row=row_idx, column=3, value=wafer.get("finalVerdict", ""))
        verdict_cell.border = border
        if verdict_cell.value == "FAIL":
            verdict_cell.font = Font(color="FF0000")
        else:
            verdict_cell.font = Font(color="00AA00")
        
        ws_wafers.cell(row=row_idx, column=4, value=f"{wafer.get('confidence', 0):.1f}%").border = border
        ws_wafers.cell(row=row_idx, column=5, value=wafer.get("severity", "")).border = border
        ws_wafers.cell(row=row_idx, column=6, value=wafer.get("detectedPattern", "")).border = border
    
    # Adjust column widths
    for col, width in [("A", 15), ("B", 25), ("C", 10), ("D", 12), ("E", 12), ("F", 18)]:
        ws_wafers.column_dimensions[col].width = width
    
    # === Trends Sheet (if data provided) ===
    if trends:
        ws_trends = wb.create_sheet("Trends")
        
        trend_headers = ["Date", "Total Wafers", "Defective", "Yield Rate"]
        for col_idx, header in enumerate(trend_headers, start=1):
            cell = ws_trends.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        for row_idx, trend in enumerate(trends, start=2):
            ws_trends.cell(row=row_idx, column=1, value=trend.get("date", "")).border = border
            ws_trends.cell(row=row_idx, column=2, value=trend.get("total_wafers", 0)).border = border
            ws_trends.cell(row=row_idx, column=3, value=trend.get("defective_wafers", 0)).border = border
            ws_trends.cell(row=row_idx, column=4, value=f"{trend.get('yield_rate', 0):.1f}%").border = border
    
    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
